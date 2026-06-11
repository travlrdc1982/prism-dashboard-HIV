"""
PRISM workbook reader — the ONE code path for study/judgments.xlsx.

After R2 the workbook carries analyst JUDGMENTS only:
    Messages        — message short-names / themes / core text labels
    SegmentMetrics  — per-segment tier, roi, highRoi, supporters,
                      activation, influence, persuadability, pre/post
                      display values + their labels

Everything identity-shaped (segment codes/names/party/population,
study meta) lives in study/study.yaml. Both consumers of workbook
judgments — extract_study.py (→ study.js / studyData.js) and
compute_core's ROI overrides — import THIS module, so the field
mapping is defined exactly once.
"""
import sys

import openpyxl


def find_col(headers, want):
    """Column index for a header name (case-insensitive, stripped)."""
    want = want.lower().strip()
    for i, h in enumerate(headers):
        if str(h or "").lower().strip() == want:
            return i
    return None


def detect_k(ws, headers):
    """
    Detect K = number of pre-post items by counting `prepost_keyN_label`
    columns in the header row that have non-empty data in the first data row.
    Returns (K, list of (pre_col, post_col, label_col, question_col, scale_col) tuples).

    POSITIONAL on purpose: each 5-column block is [pre, post, label,
    question, scale] with label at offset +2. The historical workbook has
    a header typo (items 5-7 mislabeled prepost_key4_*), so name-based
    lookup would silently map three items onto item 4's columns.
    """
    label_cols = [i for i, h in enumerate(headers)
                  if "prepost" in str(h or "").lower() and "label" in str(h or "").lower()]
    first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    item_cols = []
    for lc in label_cols:
        if first_row[lc] is None:
            continue
        # pre=lc-2, post=lc-1, question=lc+1, scale=lc+2
        item_cols.append((lc - 2, lc - 1, lc, lc + 1, lc + 2))
    return len(item_cols), item_cols


def read_messages(wb):
    """Core (token_id=0) messages: id, shortName, theme, text, spss_var."""
    ws = wb["Messages"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or (isinstance(row[0], str) and row[0].startswith("#")):
            continue
        msg_id, token_id, short_name, theme, text, spss_var = row[:6]
        if token_id == 0:
            out.append({
                "id": int(msg_id),
                "shortName": str(short_name or ""),
                "theme": str(theme or ""),
                "text": str(text or ""),
                "spss_var": str(spss_var or ""),
            })
    return out


# The judgment columns every segment row must carry. Empty cells here are
# analyst errors, not optional data — fail loudly instead of skipping.
JUDGMENT_COLS = ("tier", "roi", "supporters", "activation", "influence")


def read_segment_metrics(wb):
    """SegmentMetrics judgments, keyed by segment code.

    Returns (sm, prepost_labels, K) where sm[code] = {tier, roi, highRoi,
    supporters, activation, influence, persuadability[5], prePost{}}.
    Percent-scaled fields are round(value * 100) ints, matching both
    historical consumers (extract_hiv pct() and compute_core's ROI
    overrides used the identical transform)."""
    ws = wb["SegmentMetrics"]
    headers = [c.value for c in ws[1]]
    cols = {name: find_col(headers, name) for name in [
        "code", "roi", "highRoi", "supporters", "activation", "influence",
        "persuad_strong_support", "persuad_lean_support", "persuad_persuadable",
        "persuad_lean_oppose", "persuad_strong_oppose", "tier",
    ]}
    missing = [k for k, v in cols.items() if v is None]
    if missing:
        print(f"WARNING: missing columns in SegmentMetrics: {missing}", file=sys.stderr)

    K, item_cols = detect_k(ws, headers)

    sm = {}
    prepost_labels = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        code = str(row[0]).strip()

        empty = [c for c in JUDGMENT_COLS
                 if cols[c] is not None and row[cols[c]] is None]
        if empty:
            raise ValueError(
                f"SegmentMetrics row {code!r}: empty judgment cells {empty} — "
                f"every segment needs a complete set of analyst judgments")

        prepost = {}
        pp_metrics = []
        for i, (pre_c, post_c, lbl_c, q_c, sc_c) in enumerate(item_cols, start=1):
            pre, post, label, question, scale = (row[pre_c], row[post_c], row[lbl_c], row[q_c], row[sc_c])
            if pre is None or post is None or not label:
                continue
            key = f"item{i}"
            prepost[key] = [round(float(pre) * 100, 1), round(float(post) * 100, 1)]
            pp_metrics.append({
                "key": key, "label": str(label or ""),
                "question": str(question or ""), "scale": str(scale or ""),
            })
        if prepost_labels is None:
            prepost_labels = pp_metrics

        def pct(col):
            v = row[col] if col is not None else None
            return round(float(v or 0) * 100)

        tier_val = row[cols["tier"]] if cols["tier"] is not None and row[cols["tier"]] is not None else None
        sm[code] = {
            "roi": round(float(row[cols["roi"]] or 0), 4),
            "highRoi": pct(cols["highRoi"]),
            "supporters": pct(cols["supporters"]),
            "activation": pct(cols["activation"]),
            "influence": pct(cols["influence"]),
            "persuadability": [
                pct(cols["persuad_strong_support"]),
                pct(cols["persuad_lean_support"]),
                pct(cols["persuad_persuadable"]),
                pct(cols["persuad_lean_oppose"]),
                pct(cols["persuad_strong_oppose"]),
            ],
            "prePost": prepost,
            "tier": int(tier_val) if tier_val is not None else None,
        }
    return sm, prepost_labels, K


def load(path):
    """Open the workbook (values only)."""
    return openpyxl.load_workbook(path, data_only=True)

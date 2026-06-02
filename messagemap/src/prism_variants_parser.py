"""
PRISM variant text parser v3.

Consumes the canonical workbook structure with stable IDs and tokens:
  Column A: msg_id        (e.g., MSG_001)
  Column B: theme_label
  Column C: token         (0 = base, 1+ = proof variants)
  Column D: proof_full_text   (blank for token=0)
  Column E: proof_short_label (blank for token=0)
  Column F: core_msg_text
  Columns G-V: 16 persona variants, each header named CODE_msg_text

Each row is fully self-describing. No positional inference required.
Parser joins persona columns by segment code (from column header), not position.
"""
import openpyxl
import json
import re
from collections import OrderedDict
from typing import Dict


CODE_FROM_HEADER = re.compile(r'^([A-Z]+)_msg_text$')


def parse_variants_workbook(path: str,
                             sheet_name: str = 'Message Variants',
                             expected_segments: list = None) -> Dict:
    """Parse the canonical workbook into a structured dict ready for JSON."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]

    # Read header row 1 — find columns by name, not position
    headers = {}
    persona_cols = OrderedDict()
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=c).value
        if h is None:
            continue
        h = str(h).strip()
        headers[h] = c
        m = CODE_FROM_HEADER.match(h)
        if m:
            persona_cols[m.group(1)] = c

    # Validate required columns
    required = ['msg_id', 'theme_label', 'token', 'proof_full_text',
                'proof_short_label', 'core_msg_text']
    missing = [r for r in required if r not in headers]
    if missing:
        raise ValueError(f"Required column headers missing: {missing}")

    # CORE_msg_text is not a persona; remove if regex matched it
    persona_cols.pop('CORE', None)

    if expected_segments is not None:
        found = set(persona_cols.keys())
        expected = set(expected_segments)
        missing_segs = expected - found
        extra_segs = found - expected
        if missing_segs or extra_segs:
            raise ValueError(f"Segment mismatch in workbook headers. "
                             f"missing: {missing_segs}, extra: {extra_segs}")

    col = headers

    # Walk data rows; group by msg_id (in order of first appearance)
    msg_order = []
    by_msg = {}
    for r in range(2, ws.max_row + 1):
        msg_id = ws.cell(row=r, column=col['msg_id']).value
        if msg_id is None or not str(msg_id).strip():
            continue
        msg_id = str(msg_id).strip()
        theme = str(ws.cell(row=r, column=col['theme_label']).value or '').strip()
        token = ws.cell(row=r, column=col['token']).value
        if token is None:
            continue
        token = int(token)
        proof_full = ws.cell(row=r, column=col['proof_full_text']).value
        proof_short = ws.cell(row=r, column=col['proof_short_label']).value
        core_text = ws.cell(row=r, column=col['core_msg_text']).value

        if core_text is None:
            continue

        if msg_id not in by_msg:
            msg_order.append(msg_id)
            by_msg[msg_id] = {
                'msg_id': msg_id,
                'theme_label': theme,
                'tokens': [],
            }

        # Per-persona text for this row
        persona_text = {}
        for seg_code, ccol in persona_cols.items():
            v = ws.cell(row=r, column=ccol).value
            if v is not None and str(v).strip():
                persona_text[seg_code] = str(v).strip()

        token_record = {
            'token': token,
            'is_base': token == 0,
            'proof_short_label': (str(proof_short).strip() if proof_short else None),
            'proof_full_text':   (str(proof_full).strip()  if proof_full  else None),
            'text_core':         str(core_text).strip(),
            'text_by_persona':   persona_text,
        }
        by_msg[msg_id]['tokens'].append(token_record)

    # Sort each message's tokens by token value
    messages = []
    for mid in msg_order:
        m = by_msg[mid]
        m['tokens'].sort(key=lambda t: t['token'])
        m['n_tokens'] = len(m['tokens'])
        m['n_proofs'] = sum(1 for t in m['tokens'] if not t['is_base'])
        messages.append(m)

    return {
        'source_file': path,
        'sheet_name': sheet_name,
        'segment_codes': list(persona_cols.keys()),
        'n_messages': len(messages),
        'messages': messages,
    }


def summarize(parsed: Dict):
    print(f"Parsed: {parsed['source_file']}")
    print(f"  segments: {parsed['segment_codes']}")
    print(f"  n_messages: {parsed['n_messages']}\n")
    print(f"  {'msg_id':8s}  {'theme':28s}  {'tokens':>6s}  short_labels")
    print(f"  {'-'*8}  {'-'*28}  {'-'*6}  {'-'*60}")
    for m in parsed['messages']:
        labels = ' | '.join((t['proof_short_label'] or 'base') for t in m['tokens'])
        print(f"  {m['msg_id']:8s}  {m['theme_label']:28s}  {m['n_tokens']:>6d}  {labels[:60]}")


if __name__ == '__main__':
    path = '/mnt/user-data/uploads/Gilead_Persona-Tuned_Message_Variants_json.xlsx'
    expected = ['TSP','CEC','TC','WE','PP','HF','PFF','HHN',
                'MFL','VS','UCP','FJP','HCP','HAD','HCI','GHI']
    parsed = parse_variants_workbook(path, expected_segments=expected)
    summarize(parsed)

    out = '/home/claude/prism_variants_v3.json'
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(parsed, f, indent=2, ensure_ascii=False)
    import os
    print(f"\nWrote {out} ({os.path.getsize(out)/1024:.1f} KB)")

    # Spot-check Message 1
    msg1 = parsed['messages'][0]
    print(f"\n── Spot-check: {msg1['msg_id']} — {msg1['theme_label']} ──")
    for t in msg1['tokens']:
        fjp = t['text_by_persona'].get('FJP', '[missing]')
        label = t['proof_short_label'] or 'base'
        print(f"\n  Token {t['token']}: {label}")
        print(f"    CORE:  {t['text_core'][:110]}{'...' if len(t['text_core'])>110 else ''}")
        print(f"    FJP:   {fjp[:110]}{'...' if len(fjp)>110 else ''}")

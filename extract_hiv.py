#!/usr/bin/env python3
"""
Extract HIV study data from HIV_Study_Template.xlsx and regenerate
src/data/study.js + the HIV block of src/data/studyData.js.

K (number of pre-post items) is auto-detected from the workbook —
no hardcoded count. Tier column is found by header name (case-
insensitive, whitespace-tolerant) so header typos like "tier "
(trailing space) don't break the read.

Usage:  python extract_hiv.py
"""
import json
import sys
import openpyxl

WORKBOOK = "HIV_Study_Template.xlsx"
SEG_ORDER = ["TSP","CEC","TC","HF","PP","WE","PFF","HHN","MFL","VS",
             "UCP","FJP","HCP","HAD","HCI","GHI"]


def find_col(headers, want):
    """Find column index for a header name (case-insensitive, whitespace-stripped)."""
    want = want.lower().strip()
    for i, h in enumerate(headers):
        if str(h or "").lower().strip() == want:
            return i
    return None


def detect_k(ws, headers):
    """
    Detect K = number of pre-post items by counting `prepost_keyN_label`
    columns in the header row that have non-empty data in the first data row.
    Returns (K, list of (pre_col, post_col, label_col, question_col, scale_col) tuples).
    """
    label_cols = [i for i, h in enumerate(headers)
                  if "prepost" in str(h or "").lower() and "label" in str(h or "").lower()]
    first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    item_cols = []
    for lc in label_cols:
        if first_row[lc] is None:
            continue
        # The 5-column block is [pre, post, label, question, scale];
        # label is offset +2, so pre=lc-2, post=lc-1, question=lc+1, scale=lc+2
        item_cols.append((lc - 2, lc - 1, lc, lc + 1, lc + 2))
    return len(item_cols), item_cols


def read_meta(wb):
    meta = {}
    ws = wb["StudyMeta"]
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0] and row[1]:
            meta[str(row[0]).strip()] = row[1]
    return meta


def read_messages(wb):
    """Read core (token_id=0) messages only — Wave 1 doesn't use token variants."""
    ws = wb["Messages"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or (isinstance(row[0], str) and row[0].startswith("#")):
            continue
        msg_id, token_id, short_name, theme, text, spss_var = row[:6]
        if token_id == 0:
            out.append({
                "id": int(msg_id),
                "shortName": str(short_name or ""),
                "theme": str(theme or ""),
                "text": str(text or ""),
                "spss_var": str(spss_var or ""),
            })
    return out


def read_segment_metrics(wb):
    ws = wb["SegmentMetrics"]
    headers = [c.value for c in ws[1]]
    cols = {name: find_col(headers, name) for name in [
        "code", "name", "party", "pop", "roi", "highRoi", "supporters", "activation", "influence",
        "persuad_strong_support", "persuad_lean_support", "persuad_persuadable",
        "persuad_lean_oppose", "persuad_strong_oppose", "tier",
    ]}
    missing = [k for k, v in cols.items() if v is None]
    if missing:
        print(f"WARNING: missing columns in SegmentMetrics: {missing}", file=sys.stderr)

    K, item_cols = detect_k(ws, headers)
    print(f"Detected K={K} pre-post items. Tier column: {cols['tier'] + 1 if cols['tier'] is not None else '(not found)'}")

    sm = {}
    prepost_labels = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        code = str(row[0]).strip()

        prepost = {}
        pp_metrics = []
        for i, (pre_c, post_c, lbl_c, q_c, sc_c) in enumerate(item_cols, start=1):
            pre, post, label, question, scale = (row[pre_c], row[post_c], row[lbl_c], row[q_c], row[sc_c])
            if pre is None or post is None or not label:
                continue
            key = f"item{i}"
            prepost[key] = [round(float(pre) * 100, 1), round(float(post) * 100, 1)]
            pp_metrics.append({
                "key": key, "label": str(label or ""),
                "question": str(question or ""), "scale": str(scale or ""),
            })
        if prepost_labels is None:
            prepost_labels = pp_metrics

        def pct(col):
            v = row[col] if col is not None else None
            return round(float(v or 0) * 100)

        tier_val = row[cols["tier"]] if cols["tier"] is not None and row[cols["tier"]] is not None else None
        sm[code] = {
            "name": row[cols["name"]],
            "party": row[cols["party"]],
            "pop": round(float(row[cols["pop"]] or 0) * 100, 1),
            "roi": round(float(row[cols["roi"]] or 0), 4),
            "highRoi": pct(cols["highRoi"]),
            "supporters": pct(cols["supporters"]),
            "activation": pct(cols["activation"]),
            "influence": pct(cols["influence"]),
            "persuadability": [
                pct(cols["persuad_strong_support"]),
                pct(cols["persuad_lean_support"]),
                pct(cols["persuad_persuadable"]),
                pct(cols["persuad_lean_oppose"]),
                pct(cols["persuad_strong_oppose"]),
            ],
            "prePost": prepost,
            "tier": int(tier_val) if tier_val is not None else None,
        }
    return sm, prepost_labels, K


def js_str(s):
    return json.dumps(s, ensure_ascii=False)


def write_study_js(meta, messages, sm, prepost_labels, K):
    L = []
    L.append("// ═══════════════════════════════════════════════════════════════")
    L.append(f"// STUDY-SPECIFIC — {meta.get('topic', 'HIV Treatment & Prevention')}")
    L.append(f"// Client: {meta.get('client', 'Gilead')}")
    L.append(f"// Field date: {meta.get('field_date', '')}")
    L.append("// Generated by extract_hiv.py — do not edit by hand. Re-run the script to refresh.")
    L.append("// ═══════════════════════════════════════════════════════════════")
    L.append("")
    L.append("export const STUDY_META = {")
    L.append(f"  name: {js_str(meta.get('study_name', 'HIV'))},")
    L.append(f"  client: {js_str(meta.get('client', ''))},")
    L.append(f"  topic: {js_str(meta.get('topic', ''))},")
    L.append(f"  fieldDate: {js_str(meta.get('field_date', ''))},")
    L.append(f"  nMessages: {meta.get('n_messages', len(messages))},")
    L.append(f"  methodology: {js_str(meta.get('methodology', ''))},")
    L.append("};")
    L.append("")
    L.append(f"// K = number of pre-post items extracted from the workbook (auto-detected).")
    L.append(f"export const K_PREPOST = {K};")
    L.append("")
    L.append("// ─── MESSAGES (Wave 1: core text only, no per-segment variants or SoP yet) ───")
    L.append("export const MESSAGES = [")
    for m in messages:
        L.append(f"  {{ id:{m['id']}, shortName:{js_str(m['shortName'])}, theme:{js_str(m['theme'])}, text:{js_str(m['text'])} }},")
    L.append("];")
    L.append("")
    L.append("// SoP matrices intentionally empty for Wave 1. MessageMap renders a wave-2 placeholder.")
    L.append("export const CONTROL_SOP = [];")
    L.append("export const VARIANT_SOP = [];")
    L.append("")
    L.append("// ─── ASSIGNED TIERS (explicit, from workbook 'tier' column) ───")
    L.append("// Tiers are analyst-configured in the workbook, never derived from ROI.")
    L.append("export const ASSIGNED_TIERS = {")
    for code in SEG_ORDER:
        if code in sm and sm[code].get("tier") is not None:
            L.append(f"  {code}: {sm[code]['tier']},")
    L.append("};")
    L.append("")
    L.append("export function getAssignedTier(code) { return ASSIGNED_TIERS[code] || 3; }")
    L.append("")
    L.append("// ─── SEGMENT-LEVEL STUDY METRICS ───")
    L.append("export const STUDY_METRICS = {")
    for code in SEG_ORDER:
        if code not in sm:
            continue
        m = sm[code]
        prepost_str = ",".join(f"{k}:{json.dumps(v)}" for k, v in m["prePost"].items())
        tier = m["tier"] if m["tier"] is not None else "null"
        L.append(f"  {code}: {{ tier:{tier}, roi:{m['roi']}, highRoi:{m['highRoi']}, supporters:{m['supporters']}, activation:{m['activation']}, influence:{m['influence']}, persuadability:{json.dumps(m['persuadability'])}, prePost:{{{prepost_str}}} }},")
    L.append("};")
    L.append("")
    L.append(f"// ─── PRE/POST METRIC DEFINITIONS (K={K} items) ───")
    L.append("export const PREPOST_METRICS = [")
    for pp in prepost_labels:
        L.append(f"  {{ key:{js_str(pp['key'])}, label:{js_str(pp['label'])}, question:{js_str(pp['question'])}, scale:{js_str(pp['scale'])} }},")
    L.append("];")
    L.append("")
    L.append("// ─── THEME COLORS (intentionally empty for HIV Wave 1) ───")
    L.append("export const THEME_COLORS = {};")
    L.append("")
    L.append("export const TIER_CONFIG = {")
    L.append('  1: { bg:"#064e3b", text:"#6ee7b7", accent:"#34d399", label:"TIER 1" },')
    L.append('  2: { bg:"#854d0e", text:"#fde047", accent:"#eab308", label:"TIER 2" },')
    L.append('  3: { bg:"#991b1b", text:"#fca5a5", accent:"#ef4444", label:"TIER 3" },')
    L.append("};")
    L.append("")
    L.append("// Deprecated: tier is always analyst-configured via ASSIGNED_TIERS / getAssignedTier(code).")
    L.append("// Kept here so existing imports don't break; do not call from new code.")
    L.append("export function getTierNum(roi) { return roi >= 1.07 ? 1 : roi >= 1.00 ? 2 : 3; }")
    L.append("")
    L.append("export function getSopColor(v) {")
    L.append('  if (v >= 13) return { bg:"#065f46", t:"#6ee7b7" };')
    L.append('  if (v >= 11) return { bg:"#064e3b", t:"#6ee7b7" };')
    L.append('  if (v >= 10) return { bg:"#1a3a2a", t:"#a7f3d0" };')
    L.append('  if (v >= 9)  return { bg:"#1e293b", t:"#cbd5e1" };')
    L.append('  if (v >= 8)  return { bg:"#1a1f2e", t:"#94a3b8" };')
    L.append('  if (v >= 7)  return { bg:"#1a1520", t:"#94a3b8" };')
    L.append('  if (v >= 6)  return { bg:"#1f1318", t:"#f9a8a8" };')
    L.append('  return { bg:"#2d1215", t:"#fca5a5" };')
    L.append("}")
    L.append("")
    L.append('export const PERSUADABILITY_LABELS = ["Strong support","Lean support","Persuadable","Lean oppose","Strong oppose"];')
    L.append("")

    out = "\n".join(L)
    with open("src/data/study.js", "w") as f:
        f.write(out)
    print(f"Wrote src/data/study.js ({len(out)} bytes, K={K})")


def write_study_data_js(messages, sm, prepost_labels):
    """Replace the HIV block in studyData.js. Keeps the DATA.segments skeleton intact."""
    with open("src/data/studyData.js") as f:
        existing = f.read()
    marker = '  "HIV": {'
    idx = existing.find(marker)
    if idx < 0:
        # No HIV block yet — assume the closing has the canonical segments only.
        # Find end of segments array and inject HIV block.
        seg_end = existing.find("  ],\n};")
        if seg_end < 0:
            seg_end = existing.find("  ]\n};")
        assert seg_end > 0, "Could not locate end of segments array"
        header = existing[:seg_end + 5] + "\n"  # keep "  ],"
    else:
        header = existing[:idx]

    hiv_segments = []
    for i, code in enumerate(SEG_ORDER, start=1):
        if code not in sm:
            continue
        m = sm[code]
        hiv_segments.append({
            "id": i, "code": code, "name": m["name"], "party": m["party"],
            "pop": int(round(m["pop"])),
            "roi": m["roi"], "highRoi": m["highRoi"], "tier": m["tier"],
            "persuadability": m["persuadability"],
            "supporters": m["supporters"], "activation": m["activation"], "influence": m["influence"],
            "prePost": m["prePost"],
        })

    hiv_messages = [
        {"id": m["id"], "shortName": m["shortName"], "text": m["text"], "theme": m["theme"], "sop": []}
        for m in messages
    ]

    hiv_block = {
        "segments": hiv_segments,
        "messages": hiv_messages,
        "prePostMetrics": prepost_labels,
    }
    hiv_json = json.dumps(hiv_block, indent=2, ensure_ascii=False)
    hiv_indented = "\n".join("  " + line for line in hiv_json.split("\n"))

    new_content = header + "  \"HIV\": " + hiv_indented.lstrip() + "\n};\n\nexport default DATA;\n"
    with open("src/data/studyData.js", "w") as f:
        f.write(new_content)
    print(f"Wrote src/data/studyData.js ({len(new_content)} bytes)")


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    meta = read_meta(wb)
    messages = read_messages(wb)
    sm, prepost_labels, K = read_segment_metrics(wb)

    write_study_js(meta, messages, sm, prepost_labels, K)
    write_study_data_js(messages, sm, prepost_labels)
    print(f"Done. K={K}, {len(messages)} messages, {len(sm)} segments.")


if __name__ == "__main__":
    main()
